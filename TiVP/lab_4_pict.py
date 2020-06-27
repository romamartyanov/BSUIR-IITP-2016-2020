from openpyxl import Workbook
# from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

wb = load_workbook(filename='653502_17_2019_SFT_PICT_temp.xlsx')
ws = wb.get_sheet_by_name('653502_17_2019_SFT_PICT_temp')

wb_new = Workbook()
dest_filename = '653502_17_2019_SFT_PICT_generated.xlsx'
ws_new = wb_new.active
ws_new.title = "653502_17_2019_SFT_PICT.xlsx"

i = 1
for row in ws.iter_rows():
    print(i)
    if i == 1:
        i += 1
        continue
    cell_test_case = f'''1. Open https://www.21vek.by/mobile/
2. In {ws['A1'].value} select: {row[0].value}
3. In {ws['B1'].value} select: {row[1].value}
4. In {ws['C1'].value} select: {row[2].value}
5. In {ws['D1'].value} select: {row[3].value}
6. In {ws['E1'].value} select: {str(int(row[4].value))}
7. In {ws['F1'].value} select: {str(int(row[5].value))}
8. In {ws['G1'].value} select: {str(int(row[6].value))}
9. In {ws['H1'].value} select: {str(int(row[7].value))}
10. In {ws['I1'].value} select: {str(int(row[8].value))}
11. In {ws['J1'].value} select: {row[9].value}
12. In {ws['K1'].value} select: {row[10].value}
13. In {ws['L1'].value} select: {row[11].value}'''

    cell_result = f'''Displayed mobile phones with parameters:
{ws['A1'].value}: {row[0].value}
{ws['B1'].value}: {row[1].value}
{ws['C1'].value}: {row[2].value}
{ws['D1'].value}: {row[3].value}
{ws['E1'].value}: {str(int(row[4].value))}
{ws['F1'].value}: {str(int(row[5].value))}
{ws['G1'].value}: {str(int(row[6].value))}
{ws['H1'].value}: {str(int(row[7].value))}
{ws['I1'].value}: {str(int(row[8].value))}
{ws['J1'].value}: {row[9].value}
{ws['K1'].value}: {row[10].value}
{ws['L1'].value}: {row[11].value}'''

    ws_new[f'A{str(i)}'] = cell_test_case
    ws_new[f'B{str(i)}'] = cell_result

    i += 1

wb_new.save(filename=dest_filename)
